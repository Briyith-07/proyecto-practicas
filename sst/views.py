from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, FileResponse
from django.utils.safestring import mark_safe
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.timezone import localtime, now
from django.template.loader import render_to_string
from django.conf import settings
from django.core.mail import send_mail, EmailMultiAlternatives
from django.db.models import Count, F, Subquery, OuterRef, Max
from django.core.serializers.json import DjangoJSONEncoder
from django.db import IntegrityError
from django.urls import reverse
from django.db.models import OuterRef, Subquery, Value, CharField
from django.db.models.functions import Coalesce
from .forms import RegistrarEvidenciaCampañaForm
from .models import Campaña, EvidenciaCampaña
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.enums import TA_CENTER
from .forms import RolForm, PermisoForm
from collections import Counter
from .forms import RecursoSSTAdminForm 
from .models import RecursoSSTAdmin, RecursoSSTEmpleado
from .models import RecursoSSTEmpleado
from django.utils.dateparse import parse_date
from .forms import MensajeForm
from django.core.paginator import Paginator
from .forms import RolForm
from reportlab.lib.pagesizes import letter, landscape
from .forms import EditarEmpleadoForm
from django.db.models.functions import Trim
from django.db.models import Q


from datetime import timedelta, date
import random
import string
import json
import io
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import xlsxwriter
from .models import Perfil

from openpyxl import Workbook
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image
)
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

# Formularios
from .forms import (
    RegistroUsuarioForm, AdminCrearUsuarioForm, AdminEditarUsuarioForm,
    CampañaForm, EncuestaForm, FeedbackForm, CodigoCampañaForm,
    CampanaAsignadaForm, GrupoForm, NotificacionForm, EditarUsuarioForm
)

# Modelos
from .models import (
    Usuario, Campaña, CampanaAsignada, Feedback, CodigoCampaña,
    Calificacion, Encuesta, Grupo, Notificacion, PausaActiva, Rol, Permiso, CampañaRealizada, Mensaje
)

Usuario = get_user_model()


# Vista de inicio / landing
def inicio(request):
    return render(request, 'inicio.html', {
        'MEDIA_URL': settings.MEDIA_URL
    })

def es_admin(user):
    return user.is_authenticated and getattr(user.rol, 'nombre', '').lower() == 'admin'

# Vista de login
def login_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        usuario = authenticate(request, username=username, password=password)
        if usuario is not None:
            login(request, usuario)

            rol = usuario.rol.nombre if usuario.rol else None
            if rol == 'Administrador' or usuario.is_superuser:
                return redirect('admin_dashboard')
            elif rol == 'Empleado':
                return redirect('dashboard_empleado')
            else:
                return redirect('inicio')
        else:
            messages.error(request, 'Credenciales incorrectas')

    return render(request, 'auth/login.html')
# =====================
# LOGOUT
# =====================
def logout_view(request):
    logout(request)
    return redirect('inicio')  # Redirige a la página principal


#REESTABLECER CONTRASEÑA
def restablecer_contraseña(request, id):
    usuario = get_object_or_404(Usuario, id=id)

    if request.method == 'POST':
        nueva = request.POST.get('password')
        confirmar = request.POST.get('confirmar')

        if nueva == confirmar:
            usuario.set_password(nueva)
            usuario.codigo_temporal = None
            usuario.save()
            messages.success(request, 'Contraseña actualizada.')
            return redirect('login')
        else:
            messages.error(request, 'Las contraseñas no coinciden.')

    return render(request, 'auth/reset_password.html', {'usuario': usuario})

#enviar codigo
def enviar_codigo(request):
    if request.method == 'POST':
        correo = request.POST.get('correo')
        usuario = Usuario.objects.filter(email=correo).first()

        if not usuario:
            messages.error(request, "El correo no está registrado.")
            return redirect('enviar_codigo')

        codigo = get_random_string(length=6, allowed_chars='1234567890')
        usuario.codigo_temporal = codigo
        usuario.save()

        # Renderiza plantilla HTML personalizada
        html_content = render_to_string('auth/codigo_recuperacion.html', {
            'nombre': usuario.first_name or usuario.username,
            'codigo': codigo,
        })

        # Construye el correo con HTML
        email = EmailMultiAlternatives(
            subject="🔐 Clave Temporal de Acceso",
            body=f"Tu código temporal es: {codigo}",
            from_email="conjuntoresidencialeldorado22@gmail.com",
            to=[correo],
        )
        email.attach_alternative(html_content, "text/html")
        email.send()

        request.session['correo_verificacion'] = correo
        messages.success(request, "Hemos enviado un código a tu correo.")
        return redirect('verificar_codigo')

    return render(request, 'auth/enviar_codigo.html')
                  
#RECUPERACION DE CONTRASEÑA
def solicitar_codigo(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            usuario = Usuario.objects.get(email=email)
            codigo = str(random.randint(100000, 999999))
            usuario.codigo_temporal = codigo
            usuario.fecha_codigo = timezone.now() + timedelta(minutes=10)
            usuario.save()

            send_mail(
                subject='🔐 Código de recuperación',
                message=f'Tu código temporal es: {codigo}',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[usuario.email]
            )

            messages.success(request, 'Código enviado. Revisa tu correo.')
            return redirect('verificar_codigo')
        except Usuario.DoesNotExist:
            messages.error(request, 'Correo no registrado.')

    return render(request, 'auth/solicitar_codigo.html')

#CODIGO DE RECUPERACION
def verificar_codigo(request):
    correo = request.session.get('correo_verificacion')  # recupera correo de la sesión

    if request.method == 'POST':
        codigo_ingresado = request.POST.get('codigo')

        if not correo:
            messages.error(request, 'Correo no válido o código expirado.')
            return redirect('enviar_codigo')

        try:
            usuario = Usuario.objects.get(email=correo)
        except Usuario.DoesNotExist:
            messages.error(request, 'Correo no registrado.')
            return redirect('enviar_codigo')

        if usuario.codigo_temporal == codigo_ingresado:
            usuario.codigo_temporal = None  
            usuario.save()
            request.session['email_verificado'] = correo  
            return redirect('restablecer_contraseña', id=usuario.id)

        else:
            messages.error(request, 'Código incorrecto.')

    return render(request, 'auth/verificar_codigo.html', {'correo': correo})

#usuarios registrados
@login_required
def usuarios_registrados(request):
    usuarios = Usuario.objects.all()
    return render(request, 'usuarios_admin/usuarios_registrados.html', {'usuarios': usuarios})

#crear usuarios
@user_passes_test(lambda u: u.is_superuser)
def crear_usuario_admin(request):
    if request.method == 'POST':
        form = AdminCrearUsuarioForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('/panel/dashboard/')
    else:
        form = AdminCrearUsuarioForm()

    return render(request, 'usuarios_admin/crear_usuario_admin.html', {'form': form})

# Vista de registro de usuario
def registro(request):
    if request.method == 'POST':
        form = RegistroUsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.username = form.cleaned_data['email']  
            usuario.save()
            messages.success(request, 'Registro exitoso. Ya puedes iniciar sesión.')
            return redirect('login')
    else:
        form = RegistroUsuarioForm()
    return render(request, 'auth/registro.html', {'form': form})



# ========= VISTAS ADMIN =========

def admin_dashboard(request):
    total_usuarios = Usuario.objects.count()
    total_campañas = Campaña.objects.count()
    total_asignadas = CampanaAsignada.objects.count()
    total_grupos = Grupo.objects.count()  
    total_roles = Rol.objects.count()  
    asignaciones_detalle = CampanaAsignada.objects.select_related('campaña', 'empleado')
    
    return render(request, 'admin_dashboard.html', {
        'usuarios': total_usuarios,
        'campañas': total_campañas,
        'campanas_asignadas': total_asignadas,  
        'grupos': total_grupos,  
        'roles': total_roles,  
        'asignaciones_detalle': asignaciones_detalle,
    })
    
@user_passes_test(lambda u: u.is_superuser)
def administrar_usuarios(request):
    usuarios = Usuario.objects.all()
    return render(request, 'admin_usuarios.html', {'usuarios': usuarios})

#listar usuarios
@user_passes_test(lambda u: u.is_superuser)
def listar_usuarios(request):
    usuarios = Usuario.objects.all().select_related('perfil', 'rol')

    # Obtener filtros
    departamento = request.GET.get('departamento', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    rol = request.GET.get('rol', '').strip()
    estado = request.GET.get('estado', '').strip()

    # Aplicar filtros
    if departamento:
        usuarios = usuarios.filter(perfil__departamento__icontains=departamento)
    if ciudad:
        usuarios = usuarios.filter(perfil__ciudad__icontains=ciudad)
    if rol:
        rol_map = {'admin': 'Administrador', 'empleado': 'Empleado'}
        nombre_rol = rol_map.get(rol.lower())
        if nombre_rol:
            usuarios = usuarios.filter(rol__nombre__iexact=nombre_rol)
    if estado:
        usuarios = usuarios.filter(is_active=(estado.lower() == 'activo'))

    # Listas únicas
    departamentos = Usuario.objects.filter(perfil__departamento__isnull=False)\
                    .values_list('perfil__departamento', flat=True).distinct()
    
    # Ciudades con su departamento
    ciudades_con_departamento = Usuario.objects.filter(
        perfil__ciudad__isnull=False,
        perfil__departamento__isnull=False
    ).values_list('perfil__ciudad', 'perfil__departamento').distinct()

    context = {
        'usuarios': usuarios,
        'departamentos': departamentos,
        'ciudades_con_departamento': ciudades_con_departamento,
        'filtro_departamento': departamento,
        'filtro_ciudad': ciudad,
        'filtro_rol': rol,
        'filtro_estado': estado,
    }

    return render(request, 'usuarios_admin/listar_usuarios.html', context)

@user_passes_test(lambda u: u.is_superuser)
def crear_usuario(request):
    if request.method == 'POST':
        form = AdminCrearUsuarioForm(request.POST)
        if form.is_valid():
            try:
                # Crear usuario
                usuario = form.save(commit=False)
                usuario.date_joined = timezone.now()
                usuario.username = usuario.email
                usuario.save()

                # Crear perfil usando 'user'
                Perfil.objects.create(
                    user=usuario,
                    telefono=form.cleaned_data.get('telefono'),
                    departamento=form.cleaned_data.get('departamento'),
                    ciudad=form.cleaned_data.get('ciudad')
                )

                messages.success(request, 'Usuario y perfil creados correctamente.')
                return redirect('listar_usuarios')

            except IntegrityError as e:
                if 'cedula' in str(e):
                    messages.error(request, 'Ya existe un usuario con esa cédula.')
                elif 'email' in str(e):
                    messages.error(request, 'Ya existe un usuario con ese correo electrónico.')
                else:
                    messages.error(request, 'Ocurrió un error al crear el usuario.')
        else:
            messages.error(request, 'Corrige los errores del formulario.')
    else:
        form = AdminCrearUsuarioForm()

    return render(request, 'usuarios_admin/usuarios/crear.html', {'form': form})

# Para administradores
@login_required
def editar_usuario_admin(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)

    if request.method == "POST":
        form = AdminEditarUsuarioForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()  # Esto ya guarda usuario + perfil + rol correctamente
            messages.success(request, "Usuario actualizado correctamente.")
            return redirect('listar_usuarios')
    else:
        form = AdminEditarUsuarioForm(instance=usuario)

    context = {
        'form': form,
        'usuario': usuario,
        'departamentos': ["Cundinamarca", "Antioquia", "Valle del Cauca", "Bogotá"],
        'ciudades_por_departamento': {
            "Cundinamarca": ["Girardot", "Soacha", "Fusagasugá", "Zipaquirá", "Ricaurte"],
            "Antioquia": ["Medellín", "Envigado", "Bello", "Itagüí"],
            "Valle del Cauca": ["Cali", "Palmira", "Buenaventura", "Tuluá"],
            "Bogotá": ["Bogotá"]
        },
        'mostrar_rol': True,  # Siempre mostrar para admin
    }
    return render(request, 'usuarios_admin/editar_usuario_admin.html', context)

#inhabilitar usuario
@user_passes_test(lambda u: u.is_superuser)
def inhabilitar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.is_active = False
    usuario.save()
    messages.warning(request, f'Usuario {usuario.first_name} {usuario.last_name} inhabilitado.')
    return redirect('listar_usuarios')

#habilitar usuario
@user_passes_test(lambda u: u.is_superuser)
def habilitar_usuario(request, usuario_id):
    usuario = get_object_or_404(Usuario, id=usuario_id)
    usuario.is_active = True
    usuario.save()
    messages.success(request, 'Usuario habilitado.')
    return redirect('listar_usuarios')

#exportar
def exportar_usuarios_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    # Encabezados
    ws.append(['Nombre', 'Apellidos', 'Teléfono', 'Departamento', 'Ciudad', 'Dirección', 'Email', 'Rol', 'Estado'])

    # Datos
    for u in Usuario.objects.all().select_related('perfil', 'rol'):
        perfil = getattr(u, 'perfil', None)
        rol_nombre = u.rol.nombre if u.rol else ''
        ws.append([
            u.first_name,
            u.last_name,
            perfil.telefono if perfil else '',
            perfil.departamento if perfil else '',
            perfil.ciudad if perfil else '',
            u.direccion or '',  # <-- corregido
            u.email,
            rol_nombre,
            "Activo" if u.is_active else "Inactivo"
        ])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response


def exportar_usuarios_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="usuarios_registrados.pdf"'

    # Cambiamos a landscape para más ancho
    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(letter),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Usuarios Registrados", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [[
        'N°', 'Nombre', 'Apellidos', 'Teléfono', 'Departamento',
        'Ciudad', 'Dirección', 'Email', 'Rol', 'Estado'
    ]]

    usuarios = Usuario.objects.select_related('perfil', 'rol').all()
    for i, u in enumerate(usuarios, start=1):
        perfil = getattr(u, 'perfil', None)
        rol_nombre = u.rol.nombre if u.rol else ''
        data.append([
            str(i),
            u.first_name,
            u.last_name,
            perfil.telefono if perfil else '',
            perfil.departamento if perfil else '',
            perfil.ciudad if perfil else '',
            u.direccion or '',
            u.email,
            rol_nombre,
            "Activo" if u.is_active else "Inactivo"
        ])

    table = Table(data, repeatRows=1, hAlign='CENTER')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


def crear_campaña(request):
    if request.method == 'POST':
        form = CampañaForm(request.POST, request.FILES)
        if form.is_valid():
            campaña = form.save(commit=False)
            campaña.save()

            # Asignación a empleado individual
            empleado = form.cleaned_data.get('empleado')
            if empleado:
                CampanaAsignada.objects.create(
                    campaña=campaña,
                    empleado=empleado,
                    fecha_asignacion=timezone.now()
                )

            # Asignación a grupo
            grupo = form.cleaned_data.get('grupos')
            if grupo:
                # Añadir grupo a la relación ManyToMany
                campaña.grupos.add(grupo)

                # Asignar campaña a todos los empleados de ese grupo según su rol
                empleados_grupo = Usuario.objects.filter(grupos=grupo, rol__nombre='empleado')
                for emp in empleados_grupo:
                    # Evitar duplicados usando get_or_create
                    CampanaAsignada.objects.get_or_create(
                        campaña=campaña,
                        empleado=emp,
                        defaults={'fecha_asignacion': timezone.now()}
                    )

            messages.success(request, "Campaña creada y asignada correctamente.")
            return redirect('listar_campañas')
        else:
            messages.error(request, "Por favor corrija los errores en el formulario.")
    else:
        form = CampañaForm()

    return render(request, 'campañas/crear_campaña.html', {'form': form})

#listar campañas

@user_passes_test(lambda u: u.is_superuser)
def listar_campañas(request):
    asignacion_empleado = CampanaAsignada.objects.filter(
        campaña=OuterRef('pk')
    ).select_related('empleado')

    grupo_subquery = Grupo.objects.filter(
        campañas=OuterRef('pk')
    ).values('nombre')[:1]

    campañas = Campaña.objects.annotate(
        empleado_nombre=Subquery(asignacion_empleado.values('empleado__first_name')[:1]),
        empleado_apellido=Subquery(asignacion_empleado.values('empleado__last_name')[:1]),
        grupo_nombre=Subquery(grupo_subquery),
        asignado_a=Coalesce(
            Subquery(grupo_subquery, output_field=CharField()),
            Subquery(asignacion_empleado.values('empleado__first_name')[:1]),
            Value('No asignado'),
            output_field=CharField()
        )
    )

    # Aplicar filtros
    estado = request.GET.get('estado')
    periodicidad = request.GET.get('periodicidad')
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    if estado:
        campañas = campañas.filter(estado=estado)
    if periodicidad:
        campañas = campañas.filter(periodicidad__iexact=periodicidad)
    if fecha_desde:
        campañas = campañas.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        campañas = campañas.filter(fecha_creacion__date__lte=fecha_hasta)

    return render(request, 'campañas/listar_campañas.html', {'campañas': campañas})


#editar campaña
def editar_campaña(request, campaña_id):
    campaña = get_object_or_404(Campaña, id=campaña_id)

    if request.method == 'POST':
        form = CampañaForm(request.POST, request.FILES, instance=campaña)
        
        if form.is_valid():
            campaña = form.save(commit=False)
            campaña.save()

            # 🔹 Guardar relaciones ManyToMany correctamente
            if 'usuarios' in form.cleaned_data:
                campaña.usuarios.set(form.cleaned_data['usuarios'])

            grupo_seleccionado = form.cleaned_data.get('grupos')
            if grupo_seleccionado:
                campaña.grupos.set([grupo_seleccionado])  # pasa una lista
            else:
                campaña.grupos.clear()

            messages.success(request, "Campaña actualizada correctamente.")
            return redirect('listar_campañas')
    else:
        form = CampañaForm(instance=campaña)

    return render(request, 'campañas/editar_campaña.html', {'form': form, 'campaña': campaña})


# Eliminar campaña
def eliminar_campaña(request, id):
    campaña = get_object_or_404(Campaña, id=id)
    
    if request.method == 'POST':
        campaña.delete()
        return redirect('listar_campañas')
    
    return render(request, 'campañas/eliminar_campaña.html', {'campaña': campaña})

#exportar
def generar_grafico(datos_dict, titulo):
    fig, ax = plt.subplots()
    labels = list(datos_dict.keys())
    values = list(datos_dict.values())

    ax.bar(labels, values, color='teal')
    ax.set_title(titulo)
    ax.set_ylabel('Cantidad')
    ax.set_xlabel('Categoría')
    plt.xticks(rotation=45)
    plt.tight_layout()

    buffer = io.BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    plt.close()
    return buffer

def exportar_campañas_pdf(request):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph("Campañas Creadas", styles['Title']))
    elements.append(Spacer(1, 12))

    # Cabecera tabla
    data = [["ID", "Código", "Nombre", "Detalle", "Estado", "Periodicidad", "Fecha"]]

    campañas = Campaña.objects.all()

    # Diccionarios de conteo
    estado_count = {e[0]: 0 for e in Campaña.ESTADOS}
    periodicidad_count = {p[0]: 0 for p in Campaña.PERIODICIDADES}

    for c in campañas:
        data.append([
            str(c.id),
            c.codigo.codigo,
            c.codigo.nombre,
            c.detalle,
            c.estado,
            c.periodicidad,
            localtime(c.fecha_creacion).strftime("%Y-%m-%d %H:%M")
        ])
        if c.estado in estado_count:
            estado_count[c.estado] += 1
        if c.periodicidad in periodicidad_count:
            periodicidad_count[c.periodicidad] += 1

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2B547E")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
        ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
    ]))

    elements.append(table)
    elements.append(PageBreak())

    # Gráficos
    grafico_estado = generar_grafico(estado_count, "Distribución por Estado de Campañas")
    elements.append(Paragraph("Gráfico: Estado de Campañas", styles['Heading2']))
    elements.append(Spacer(1, 10))
    elements.append(Image(grafico_estado, width=450, height=200))
    elements.append(Spacer(1, 20))
    grafico_periodicidad = generar_grafico(periodicidad_count, "Distribución por Periodicidad")
    elements.append(Paragraph("Gráfico: Periodicidad de Campañas", styles['Heading2']))
    elements.append(Spacer(1, 10))
    elements.append(Image(grafico_periodicidad, width=450, height=200))

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={"Content-Disposition": 'inline; filename="campañas_creadas.pdf"'}
    )
   
def exportar_campañas_excel(request):
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output, {'in_memory': True})

    # Hoja 1: Tabla de campañas
    ws1 = workbook.add_worksheet("Campañas")
    headers = ["ID", "Código", "Nombre", "Detalle", "Estado", "Periodicidad", "Fecha"]
    for col_num, header in enumerate(headers):
        ws1.write(0, col_num, header)

    campañas = Campaña.objects.all()

    # Diccionarios de conteo
    estado_count = {e[0]: 0 for e in Campaña.ESTADOS}
    periodicidad_count = {p[0]: 0 for p in Campaña.PERIODICIDADES}

    for row_num, c in enumerate(campañas, start=1):
        ws1.write(row_num, 0, c.id)
        ws1.write(row_num, 1, c.codigo.codigo)
        ws1.write(row_num, 2, c.codigo.nombre)
        ws1.write(row_num, 3, c.detalle)
        ws1.write(row_num, 4, c.estado)
        ws1.write(row_num, 5, c.periodicidad)
        ws1.write(row_num, 6, c.fecha_creacion.strftime("%Y-%m-%d %H:%M"))

        if c.estado in estado_count:
            estado_count[c.estado] += 1
        if c.periodicidad in periodicidad_count:
            periodicidad_count[c.periodicidad] += 1

    # Hoja 2: Gráficas
    ws2 = workbook.add_worksheet("Gráficas")

    # Estado
    ws2.write(0, 0, "Estado")
    ws2.write(0, 1, "Cantidad")
    for i, (estado, count) in enumerate(estado_count.items(), start=1):
        ws2.write(i, 0, estado)
        ws2.write(i, 1, count)

    chart1 = workbook.add_chart({'type': 'column'})
    chart1.add_series({
        'name': 'Estado',
        'categories': ['Gráficas', 1, 0, len(estado_count), 0],
        'values':     ['Gráficas', 1, 1, len(estado_count), 1],
    })
    chart1.set_title({'name': 'Distribución por Estado'})
    chart1.set_style(10)
    ws2.insert_chart('D2', chart1, {'x_offset': 25, 'y_offset': 10})

    # Periodicidad
    start_row = len(estado_count) + 3
    ws2.write(start_row, 0, "Periodicidad")
    ws2.write(start_row, 1, "Cantidad")
    for i, (p, count) in enumerate(periodicidad_count.items(), start=start_row + 1):
        ws2.write(i, 0, p)
        ws2.write(i, 1, count)

    chart2 = workbook.add_chart({'type': 'column'})
    chart2.add_series({
        'name': 'Periodicidad',
        'categories': ['Gráficas', start_row + 1, 0, i, 0],
        'values':     ['Gráficas', start_row + 1, 1, i, 1],
    })
    chart2.set_title({'name': 'Distribución por Periodicidad'})
    chart2.set_style(11)
    ws2.insert_chart('D20', chart2, {'x_offset': 25, 'y_offset': 10})

    workbook.close()
    output.seek(0)

    return HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=campañas_creadas.xlsx'}
    )
    
#crear codigos
def crear_codigo(request):
    if request.method == 'POST':
        form = CodigoCampañaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_codigos')  
    else:
        form = CodigoCampañaForm()
    
    return render(request, 'codigos/crear_codigo.html', {'form': form})

#editar codigo
def editar_codigo(request, id):
    codigo = get_object_or_404(CodigoCampaña, id=id)

    if request.method == 'POST':
        form = CodigoCampañaForm(request.POST, instance=codigo)
        if form.is_valid():
            form.save()
            return redirect('listar_codigos')  
    else:
        form = CodigoCampañaForm(instance=codigo)

    return render(request, 'codigos/editar_codigo.html', {'form': form})

#eliminar codigo
def eliminar_codigo(request, id):
    codigo = get_object_or_404(CodigoCampaña, id=id)
    
    if request.method == 'POST':
        codigo.delete()
        return redirect('listar_codigos')

#listar codigos
def listar_codigos(request):
    codigos = CodigoCampaña.objects.all()
    return render(request, 'codigos/listar_codigos.html', {'codigos': codigos})

def campañas_completas(request):
    campañas = Campaña.objects.all()

    datos_tarjetas = [
        {'titulo': 'Campañas Creadas', 'total': campañas.count(), 'color': 'primary'},
    ]

    return render(request, 'estadisticas/estadisticas_campañas.html', {
        'datos_tarjetas': datos_tarjetas,
    })

def campañas_resumen(request):
    campañas = Campaña.objects.all()
    asignadas = CampanaAsignada.objects.select_related('campaña', 'empleado').all()

    # JSON para gráficos
    campañas_json = json.dumps(
        [{'nombre': c.codigo.nombre, 'estado': c.estado, 'periodicidad': c.periodicidad,
          'cantidad': c.campanaasignada_set.count()} for c in campañas],
        cls=DjangoJSONEncoder
    )

    datos_tarjetas = [
        {'titulo': 'Campañas Creadas', 'total': campañas.count(), 'color': 'primary'},
    ]

    return render(request, 'estadisticas/campañas_resumen.html', {
        'campañas': campañas,
        'asignadas': asignadas,
        'datos_tarjetas': datos_tarjetas,
        'campañas_json': campañas_json,
    })

def campañas_resumen(request):
    campañas = Campaña.objects.all()

    datos_tarjetas = [
        {'titulo': 'Campañas Creadas', 'total': campañas.count(), 'color': 'primary'},
    ]

    
    campañas_json = json.dumps(
        [{'nombre': c.codigo.nombre, 'periodicidad': c.periodicidad, 'cantidad': c.campanaasignada_set.count()} for c in campañas],
        cls=DjangoJSONEncoder
    )

    return render(request, 'estadisticas/campañas_resumen.html', {
        'campañas': campañas,
        'datos_tarjetas': datos_tarjetas, 
        'campañas_json': campañas_json,
    })

def generar_reportes(request):
    campañas = Campaña.objects.all()
    grupos = Grupo.objects.all()

    return render(request, "reportes/generar_reportes.html", {
        "campañas": campañas,
        "grupos": grupos,
    })

def exportar_reportes(request):
    tipo = request.GET.get("tipo")
    formato = request.GET.get("formato")

    # Subfiltros
    estado_usuario = request.GET.get("estado_usuario")
    estado_campaña = request.GET.get("estado_campaña")
    periodicidad = request.GET.get("periodicidad")
    campaña_id = request.GET.get("campa%C3%B1a_id") or request.GET.get("campaña_id")
    grupo_id = request.GET.get("grupo_id")
    nombre_campaña = request.GET.get("nombre_campaña")
    rol_id = request.GET.get("rol_id")  

    # -------------------
    # FILTRO PRINCIPAL
    # -------------------
    if tipo == "usuarios":
        qs = Usuario.objects.all()
        if estado_usuario == "habilitados":
            qs = qs.filter(is_active=True)
        elif estado_usuario == "inhabilitados":
            qs = qs.filter(is_active=False)
        nombre_reporte = "usuarios"

    elif tipo == "campañas":
        qs = Campaña.objects.all()
        if estado_campaña:
            qs = qs.filter(estado__iexact=estado_campaña)
        periodicidad_campaña = request.GET.get("periodicidad_campaña")
        if periodicidad_campaña:
            qs = qs.filter(periodicidad__iexact=periodicidad_campaña)
        if campaña_id:
            qs = qs.filter(id=campaña_id)
        nombre_reporte = "campañas"

    elif tipo == "grupos":
        qs = Grupo.objects.all()
        nombre_reporte = "grupos"

    elif tipo == "roles":   
        qs = Rol.objects.all()
        if rol_id:
            qs = qs.filter(id=rol_id)
        nombre_reporte = "roles"

    else:
        return HttpResponse("Tipo inválido.")

    # -------------------
    # EXPORTAR A EXCEL
    # -------------------
    if formato == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # ------------------- USUARIOS
        if tipo == "usuarios":
            ws.append([
                "ID", "Nombre", "Apellidos", "Cédula", "Teléfono",
                "Departamento", "Ciudad", "Dirección", "Email", "Rol", "Estado"
            ])
            qs = qs.order_by("id")
            for u in qs:
                rol = u.rol.nombre.capitalize() if hasattr(u, 'rol') and u.rol else "Empleado"
                estado = "Activo" if u.is_active else "Inactivo"

                ws.append([
                    u.id, u.first_name, u.last_name, u.cedula, u.telefono,
                    u.departamento, u.ciudad, u.direccion, u.email, rol, estado
                ])

        # ------------------- CAMPAÑAS
        elif tipo == "campañas":
            ws.append(["ID", "Código", "Nombre", "Estado", "Periodicidad"])
            for c in qs.order_by("id"):
                ws.append([
                    c.id,
                    c.codigo.codigo if c.codigo else "",
                    c.codigo.nombre if c.codigo else "",
                    c.estado,
                    c.periodicidad
                ])

        # ------------------- GRUPOS
        elif tipo == "grupos":
            ws.append(["ID", "Nombre", "Descripción", "Usuarios"])
            for g in qs.order_by("id"):
                usuarios = ", ".join([f"{u.first_name} {u.last_name}" for u in g.usuarios.all()])
                ws.append([g.id, g.nombre, g.descripcion or "", usuarios])

        # ------------------- ROLES  
        elif tipo == "roles":
            ws.append(["ID", "Nombre", "Permisos"])
            for r in qs.order_by("id"):
                permisos = ", ".join([p.nombre for p in r.permisos.all()])
                ws.append([r.id, r.nombre, permisos])

        # Render Excel
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre_reporte}.xlsx"'
        wb.save(response)
        return response

    # -------------------
    # EXPORTAR A PDF
    # -------------------
    elif formato == "pdf":
        from reportlab.lib.pagesizes import landscape, A4

        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = f'attachment; filename="{nombre_reporte}.pdf"'

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=15,
            rightMargin=15,
            topMargin=15,
            bottomMargin=15
        )

        elements = []
        styles = getSampleStyleSheet()

        header_style = ParagraphStyle(
            name='header_style',
            parent=styles['Heading2'],
            alignment=TA_CENTER,
            fontSize=16,
            leading=20
        )
        cell_style = ParagraphStyle(
            name='cell_style',
            parent=styles['Normal'],
            fontSize=9,
            leading=11,
            alignment=TA_CENTER
        )

        elements.append(Paragraph(f"Reporte: {nombre_reporte.upper()}", header_style))
        elements.append(Spacer(1, 12))


        # ------------------- USUARIOS
        if tipo == "usuarios":
            data = [[
                Paragraph("ID", cell_style),
                Paragraph("Nombre", cell_style),
                Paragraph("Apellidos", cell_style),
                Paragraph("Cédula", cell_style),
                Paragraph("Teléfono", cell_style),
                Paragraph("Departamento", cell_style),
                Paragraph("Ciudad", cell_style),
                Paragraph("Dirección", cell_style),
                Paragraph("Email", cell_style),
                Paragraph("Rol", cell_style),
                Paragraph("Estado", cell_style),
            ]]
            for u in qs:
                data.append([
                    Paragraph(str(u.id), cell_style),
                    Paragraph(u.first_name or "", cell_style),
                    Paragraph(u.last_name or "", cell_style),
                    Paragraph(str(u.cedula or ""), cell_style),
                    Paragraph(str(u.telefono or ""), cell_style),
                    Paragraph(u.departamento or "", cell_style),
                    Paragraph(u.ciudad or "", cell_style),
                    Paragraph(u.direccion or "", cell_style),
                    Paragraph(u.email or "", cell_style),
                    Paragraph(u.rol.nombre if u.rol else "Empleado", cell_style),
                    Paragraph("Activo" if u.is_active else "Inactivo", cell_style),
                ])
            col_widths = [30, 60, 80, 70, 70, 90, 90, 100, 100, 70, 50]

        # ------------------- CAMPAÑAS
        elif tipo == "campañas":
            data = [[
                Paragraph("ID", cell_style),
                Paragraph("Código", cell_style),
                Paragraph("Nombre", cell_style),
                Paragraph("Estado", cell_style),
                Paragraph("Periodicidad", cell_style),
            ]]
            for c in qs:
                data.append([
                    Paragraph(str(c.id), cell_style),
                    Paragraph(c.codigo.codigo if c.codigo else "", cell_style),
                    Paragraph(c.codigo.nombre if c.codigo else "", cell_style),
                    Paragraph(c.estado or "", cell_style),
                    Paragraph(c.periodicidad or "", cell_style),
                ])
            col_widths = [30, 60, 250, 70]

        # ------------------- GRUPOS
        elif tipo == "grupos":
            data = [[
                Paragraph("ID", cell_style),
                Paragraph("Nombre", cell_style),
                Paragraph("Descripción", cell_style),
                Paragraph("Usuarios", cell_style),
            ]]
            for g in qs:
                usuarios = ", ".join([f"{u.first_name} {u.last_name}" for u in g.usuarios.all()])
                data.append([
                    Paragraph(str(g.id), cell_style),
                    Paragraph(g.nombre, cell_style),
                    Paragraph(g.descripcion or "", cell_style),
                    Paragraph(usuarios, cell_style),
                ])
            col_widths = [30, 150, 200, 250]

        # ------------------- ROLES  
        elif tipo == "roles":
            data = [[
                Paragraph("ID", cell_style),
                Paragraph("Nombre", cell_style),
                Paragraph("Permisos", cell_style),
            ]]
            for r in qs:
                permisos = ", ".join([p.nombre for p in r.permisos.all()])
                data.append([
                    Paragraph(str(r.id), cell_style),
                    Paragraph(r.nombre, cell_style),
                    Paragraph(permisos, cell_style),
                ])
            col_widths = [30, 150, 350]

        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#91E5FC")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,1), (0,-1), 'CENTER'),
            ('ALIGN', (1,1), (1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ]))

        elements.append(table)
        doc.build(elements)
        return response

    else:
        return HttpResponse("Formato inválido.")


def generar_reportes(request):
    return render(request, 'estadisticas/descargar_reportes.html')

# === GRUPOS ===
@user_passes_test(lambda u: u.is_superuser)
def listar_grupos(request):
    nombre_filtro = request.GET.get('nombre', '')
    usuario_filtro = request.GET.get('usuario', '')

    grupos = Grupo.objects.all()

    if nombre_filtro:
        grupos = grupos.filter(nombre=nombre_filtro)

    if usuario_filtro:
        grupos = grupos.filter(usuarios__id=usuario_filtro)

    nombres_grupos = Grupo.objects.values_list('nombre', flat=True).distinct()
    usuarios = Usuario.objects.all()

    return render(request, 'Grupos/listar_grupos.html', {
        'grupos': grupos,
        'nombres_grupos': nombres_grupos,
        'usuarios': usuarios,
        'nombre_filtro': nombre_filtro,
        'usuario_filtro': usuario_filtro,
    })

@user_passes_test(lambda u: u.is_superuser)
def crear_grupo(request):
    if request.method == 'POST':
        form = GrupoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_grupos') 
    else:
        form = GrupoForm()
    return render(request, 'grupos/crear_grupo.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def editar_grupo(request, id):
    grupo = get_object_or_404(Grupo, id=id)
    if request.method == 'POST':
        form = GrupoForm(request.POST, instance=grupo)
        if form.is_valid():
            form.save()
            messages.success(request, "Grupo actualizado.")
            return redirect('listar_grupos')
    else:
        form = GrupoForm(instance=grupo)
    return render(request, 'grupos/editar_grupo.html', {'form': form, 'grupo': grupo})

@user_passes_test(lambda u: u.is_superuser)

def eliminar_grupo(request, id):
    grupo = get_object_or_404(Grupo, id=id)
    if request.method == 'POST':
        grupo.delete()
        messages.warning(request, "Grupo eliminado.")
        return redirect('listar_grupos')
    return render(request, 'grupos/eliminar_grupo.html', {'grupo': grupo})


#Notificaciones admin#
def notificaciones_json(request):
    notificaciones = Notificacion.objects.filter(
        usuario=request.user,
        tipo="web"
    ).order_by("-fecha_envio")[:10]  

    data = {
        "count": notificaciones.filter(abierta=False).count(),
        "items": [
            {
                "id": n.id,
                "titulo": n.titulo,
                "mensaje": n.mensaje,
                "fecha": localtime(n.fecha_envio).strftime("%d/%m/%Y %H:%M"),
                "abierta": n.abierta,
            }
            for n in notificaciones
        ]
    }
    return JsonResponse(data)

@login_required
def marcar_notificacion_leida(request, pk):
    if request.method == "POST":
        notificacion = get_object_or_404(Notificacion, id=pk, usuario=request.user)
        if not notificacion.abierta:
            notificacion.abierta = True
            notificacion.fecha_apertura = timezone.now()
            notificacion.save(update_fields=["abierta", "fecha_apertura"])
        return JsonResponse({"status": "ok"})
    return JsonResponse({"status": "error", "message": "Método no permitido"}, status=405)

def api_notificaciones(request):
    usuario = request.user
    notificaciones = Notificacion.objects.filter(usuario=usuario, abierta=False).order_by('-fecha_envio')[:10]
    
    data = [
        {
            "id": n.id,
            "titulo": n.titulo,
            "mensaje": n.mensaje,
            "fecha_envio": n.fecha_envio.strftime("%d/%m/%Y %H:%M"),
        }
        for n in notificaciones
    ]
    return JsonResponse(data, safe=False)

def listar_notificaciones(request):

    # ---------------------------
    # Capturar filtros
    # ---------------------------
    titulo_filtro = request.GET.get('titulo', '')
    usuario_filtro = request.GET.get('usuario', '')
    abierta_filtro = request.GET.get('abierta', '')

    # Query base
    notificaciones = Notificacion.objects.all().select_related("campaña", "usuario").order_by("-id")

    # ---------------------------
    # Aplicar filtros
    # ---------------------------
    if titulo_filtro:
        notificaciones = notificaciones.filter(titulo=titulo_filtro)

    if usuario_filtro:
        notificaciones = notificaciones.filter(usuario_id=usuario_filtro)

    if abierta_filtro != "":
        notificaciones = notificaciones.filter(abierta=(abierta_filtro == "1"))  # 1=Sí, 0=No

    # ---------------------------
    # Paginación 10 por página
    # ---------------------------
    paginator = Paginator(notificaciones, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ---------------------------
    # Preparar listas para filtros
    # ---------------------------
    titulos = Notificacion.objects.values_list("titulo", flat=True).distinct()
    usuarios = Usuario.objects.all().order_by("first_name")

    # ---------------------------
    # Render
    # ---------------------------
    return render(
        request,
        "Notificaciones/listar_notificaciones.html",
        {
            "notificaciones": page_obj,
            "titulos": titulos,
            "usuarios": usuarios,
            "titulo_filtro": titulo_filtro,
            "usuario_filtro": usuario_filtro,
            "abierta_filtro": abierta_filtro,
        }
    )

def crear_notificacion(request):
    if request.method == "POST":
        form = NotificacionForm(request.POST)
        if form.is_valid():
            notificacion = form.save(commit=False)
            notificacion.enviada = True
            notificacion.save()
            messages.success(request, "✅ Notificación creada y enviada con éxito.")
            return redirect("listar_notificaciones")
        else:
            messages.error(request, "❌ Error al crear la notificación. Revisa los campos.")
    else:
        form = NotificacionForm()

    # 🔥 Obtener campañas con código y nombre
    campañas = Campaña.objects.all().values(
        'id',
        'nombre',
        'codigo__codigo',
        'detalle',
        'estado',
        'periodicidad',
        'multimedia',
        'fecha_creacion'
    )

    campañas_info = []
    for c in campañas:
        # Preparar URL de multimedia
        c['multimedia'] = f"{settings.MEDIA_URL}{c['multimedia']}" if c['multimedia'] else ""
        # Combinar código + nombre para mostrar en frontend
        c['codigo_nombre'] = f"{c['codigo__codigo']} - {c['nombre']}"
        campañas_info.append(c)

    # Obtener usuarios
    usuarios = Usuario.objects.all().values('id', 'first_name', 'last_name', 'cedula')

    return render(request, 'Notificaciones/crear_notificaciones.html', {
        'form': form,
        'campañas_info': campañas_info,
        'usuarios_info': list(usuarios),
    })

# Vista para administradores
def detalle_notificacion_admin(request, pk):
    notificacion = get_object_or_404(Notificacion, pk=pk)

    return render(request, "Notificaciones/detalle_notificacion_admin.html", {
        "notificacion": notificacion
    })


#editar notificaciones
def editar_notificacion(request, pk):
    notificacion = get_object_or_404(Notificacion, pk=pk)
    if request.method == "POST":
        form = NotificacionForm(request.POST, instance=notificacion)
        if form.is_valid():
            form.save()
            return redirect("listar_notificaciones")
    else:
        form = NotificacionForm(instance=notificacion)
    return render(
        request,
        "Notificaciones/editar_notificaciones.html",
        {"form": form, "notificacion": notificacion}
    )


#eliminar notificaciones
def eliminar_notificacion(request, pk):
    notificacion = get_object_or_404(Notificacion, pk=pk)
    if request.method == "POST":
        notificacion.delete()
        return redirect("listar_notificaciones")
    return render(request, "Notificaciones/eliminar_notificaciones.html", {"notificacion": notificacion})


#notificacion empleado
@login_required
def listar_notificaciones_empleado(request):
    notificaciones = Notificacion.objects.filter(usuario=request.user)

    estado = request.GET.get('estado', '')  
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    if estado == 'leida':
        notificaciones = notificaciones.filter(abierta=True)
    elif estado == 'noleida':
        notificaciones = notificaciones.filter(abierta=False)

    if fecha_desde:
        notificaciones = notificaciones.filter(fecha_envio__date__gte=fecha_desde)
    if fecha_hasta:
        notificaciones = notificaciones.filter(fecha_envio__date__lte=fecha_hasta)

    notificaciones = notificaciones.order_by('-fecha_envio')

    # --- PAGINACIÓN ----
    paginator = Paginator(notificaciones, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "Notificaciones_empleados/listar_notificaciones_empleado.html", {
        "notificaciones": page_obj,   # <<< IMPORTANTE
        "page_obj": page_obj,         # <<< NECESARIO PARA TU HTML
        "estado": estado,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
    })

def campanas_admin(request):
    asignadas = CampanaAsignada.objects.select_related('campaña', 'empleado')

    # Filtros desde GET
    empleado_id = request.GET.get('empleado')
    estado = request.GET.get('estado')
    periodicidad = request.GET.get('periodicidad')

    if empleado_id:
        asignadas = asignadas.filter(empleado__id=empleado_id)
    if estado:
        asignadas = asignadas.filter(campaña__estado=estado)
    if periodicidad:
        asignadas = asignadas.filter(campaña__periodicidad=periodicidad)

    # Determinar estado_actual considerando evidencia
    for a in asignadas:
        evidencia = EvidenciaCampaña.objects.filter(campaña=a.campaña, empleado=a.empleado).first()
        if evidencia:
            if a.campaña.estado == 'por_aprobacion':
                a.estado_actual = 'En revisión'
            elif a.campaña.estado == 'finalizada':
                a.estado_actual = 'Finalizada'
            else:
                a.estado_actual = a.campaña.estado
        else:
            a.estado_actual = a.campaña.estado

   
    empleados = Usuario.objects.filter(rol__nombre__iexact='Empleado')

    estados = [
        ('activa', 'Activa'),
        ('pausada', 'Pausada'),
        ('por_aprobacion', 'Por aprobación'),
        ('finalizada', 'Finalizada'),
        ('rechazada', 'Rechazada'),
    ]
    periodicidades = [
        ('Diaria', 'Diaria'),
        ('Semanal', 'Semanal'),
        ('Mensual', 'Mensual'),
    ]

    context = {
        'asignadas': asignadas,
        'empleados': empleados,
        'estados': estados,
        'periodicidades': periodicidades,
        'filtro_empleado': empleado_id or '',
        'filtro_estado': estado or '',
        'filtro_periodicidad': periodicidad or '',
    }

    return render(request, 'campanas_admin/campanas_admin.html', context)


def detalle_campana_admin(request, id):
    asignacion = get_object_or_404(CampanaAsignada, id=id)
    evidencia = EvidenciaCampaña.objects.filter(
        campaña=asignacion.campaña,
        empleado=asignacion.empleado
    ).first()

    if request.method == 'POST':
        accion = request.POST.get('accion')
        if evidencia:
            if accion == 'aprobar':
                asignacion.campaña.estado = 'finalizada'
            elif accion == 'rechazar':
                asignacion.campaña.estado = 'rechazada'
            asignacion.campaña.save()
        return redirect('campanas_admin')

    return render(
        request,
        'campanas_admin/detalle_campana_admin.html',
        {
            'asignacion': asignacion,
            'campaña': asignacion.campaña,
            'evidencia': evidencia,
        }
    )


def aprobar_campaña(request, id):
    asignada = get_object_or_404(CampanaAsignada, id=id)
    asignada.campaña.estado = 'finalizada'
    asignada.campaña.save()
    messages.success(request, f"La campaña '{asignada.campaña.nombre}' fue aprobada correctamente.")
    return redirect('campanas_admin')


def rechazar_campaña(request, id):
    asignada = get_object_or_404(CampanaAsignada, id=id)
    asignada.campaña.estado = 'activa'
    asignada.campaña.save()
    messages.warning(request, f"La campaña '{asignada.campaña.nombre}' fue rechazada.")
    return redirect('campanas_admin')

# Lista de roles
def listar_roles(request):
    roles = Rol.objects.prefetch_related('permisos').all()  # Trae los permisos de cada rol
    return render(request, 'roles/listar_roles.html', {'roles': roles})

def crear_rol(request):
    if request.method == 'POST':
        form = RolForm(request.POST)
        if form.is_valid():
            rol = form.save(commit=False)
            rol.save()
            form.save_m2m()  # Muy importante: guarda los permisos ManyToMany
            return redirect('listar_roles')
    else:
        form = RolForm()

    categorias = {
        "Usuarios": ["Usuarios Registrados","Ver usuarios","Crear usuario","Modificar usuario","Habilitar usuario","Inhabilitar usuario"],
        "Campañas": ["Crear campaña","Modificar campaña","Eliminar campaña","Ver campañas"],
        "Estadísticas": ["Campañas Creadas (para estadísticas)","Descargar reportes"],
        "Notificaciones": ["Crear notificación","Modificar notificación","Eliminar notificación","Ver notificaciones"],
        "Grupos": ["Crear grupo","Modificar grupo","Eliminar grupo","Ver grupos"],
        "Roles y Permisos": ["Crear rol","Modificar rol","Asignar permisos","Ver roles"],
        "Recursos SST": ["Subir recurso","Modificar recurso","Eliminar recurso","Ver recurso"],
        "Mensajes": ["Enviar mensaje","Modificar mensaje","Eliminar mensaje","Ver mensajes"],
        "Auditoría": ["Ver Logs"]
    }

    # Pasa el queryset completo de permisos al template
    permisos = Permiso.objects.all()

    return render(request, 'roles/crear_rol.html', {'form': form, 'categorias': categorias, 'permisos': permisos})

# Editar rol
def editar_rol(request, pk):
    rol = get_object_or_404(Rol, pk=pk)
    if request.method == "POST":
        form = RolForm(request.POST, instance=rol)
        if form.is_valid():
            form.save()
            return redirect('listar_roles')
    else:
        form = RolForm(instance=rol)
    return render(request, 'roles/editar_rol.html', {'form': form, 'rol': rol})

# Eliminar rol
def eliminar_rol(request, pk):
    rol = get_object_or_404(Rol, pk=pk)
    if request.method == "POST":
        rol.delete()
        return redirect('listar_roles')
    return render(request, 'roles/eliminar_rol.html', {'rol': rol})

#Recursos Admin#--------------
@login_required
def recursos_sst_admin(request):
    recursos = RecursoSSTAdmin.objects.all()

    # === FILTROS ===
    titulo = request.GET.get("titulo")
    tipo = request.GET.get("tipo")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if titulo:
        recursos = recursos.filter(titulo__icontains=titulo)

    if tipo:
        recursos = recursos.filter(tipo=tipo)

    if fecha_desde:
        recursos = recursos.filter(fecha_subida__date__gte=fecha_desde)

    if fecha_hasta:
        recursos = recursos.filter(fecha_subida__date__lte=fecha_hasta)

    # === CREAR RECURSO ===
    if request.method == 'POST':
        form = RecursoSSTAdminForm(request.POST, request.FILES)
        if form.is_valid():
            nuevo_recurso = form.save()

            # Crear copia visible para empleados
            RecursoSSTEmpleado.objects.create(recurso=nuevo_recurso, visible=True)

            messages.success(request, "Recurso creado y visible para empleados.")
            return redirect('recursos_sst_admin')

    else:
        form = RecursoSSTAdminForm()

    return render(request, 'Recursos_admin/recursos_sst_admin.html', {
        'recursos': recursos,
        'form': form,
    })


@login_required
def crear_recurso(request):
    if request.method == 'POST':
        form = RecursoSSTAdminForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Recurso creado correctamente")
            return redirect('recursos_sst_admin')
    else:
        form = RecursoSSTAdminForm()

    return render(request, 'Recursos_admin/crear_recurso.html', {'form': form})


@login_required
def editar_recurso(request, pk):
    recurso = get_object_or_404(RecursoSSTAdmin, pk=pk)
    if request.method == 'POST':
        form = RecursoSSTAdminForm(request.POST, request.FILES, instance=recurso)
        if form.is_valid():
            form.save()
            messages.success(request, "Recurso actualizado correctamente")
            return redirect('recursos_sst_admin')
    else:
        form = RecursoSSTAdminForm(instance=recurso)

    return render(request, 'Recursos_admin/editar_recurso.html', {'form': form, 'recurso': recurso})

@login_required
def eliminar_recurso(request, pk):
    recurso = get_object_or_404(RecursoSSTAdmin, pk=pk)
    if request.method == 'POST':
        recurso.delete()
        messages.success(request, "Recurso eliminado correctamente")
        return redirect('recursos_sst_admin')

    return render(request, 'Recursos_admin/eliminar_recurso.html', {'recurso': recurso})
 
 #Mensajes
@login_required
def panel_mensajes(request):
    mensajes = Mensaje.objects.all()

    titulo = request.GET.get('titulo', '')
    fecha_evento_desde = request.GET.get('fecha_evento_desde', '')
    fecha_evento_hasta = request.GET.get('fecha_evento_hasta', '')

    if titulo:
        mensajes = mensajes.filter(titulo=titulo)
    if fecha_evento_desde:
        mensajes = mensajes.filter(fecha_evento__date__gte=fecha_evento_desde)
    if fecha_evento_hasta:
        mensajes = mensajes.filter(fecha_evento__date__lte=fecha_evento_hasta)

    titulos_disponibles = Mensaje.objects.values_list('titulo', flat=True).distinct()

    paginator = Paginator(mensajes, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "mensajes_admin/panel_mensajes.html", {
        "mensajes": page_obj.object_list,
        "page_obj": page_obj,
        "titulos_disponibles": titulos_disponibles,
        "filtros": {
            "titulo": titulo,
            "fecha_evento_desde": fecha_evento_desde,
            "fecha_evento_hasta": fecha_evento_hasta,
        }
    })

@login_required
def crear_mensaje(request):
    if request.method == "POST":
        form = MensajeForm(request.POST)
        if form.is_valid():
            mensaje = form.save()

            # Crear notificaciones web
            empleados = Usuario.objects.filter(rol__nombre="empleado")  
            for empleado in empleados:
                # Construir mensaje con fecha y hora del evento
                mensaje_web = mensaje.contenido
                if mensaje.fecha_evento:
                    mensaje_web += f"\n📅 Fecha del evento: {mensaje.fecha_evento.strftime('%d/%m/%Y %H:%M')}"

                Notificacion.objects.create(
                    usuario=empleado,
                    remitente=request.user,  
                    titulo=f"Nuevo mensaje: {mensaje.titulo}",
                    mensaje=mensaje_web,
                    tipo="web",
                    enviada=True
                )

            messages.success(request, "Mensaje creado y enviado correctamente.")
            return redirect("panel_mensajes")
    else:
        form = MensajeForm()

    return render(request, "mensajes_admin/crear_mensaje.html", {"form": form})


def editar_mensaje(request, id):
    mensaje = get_object_or_404(Mensaje, id=id)

    if not request.user.is_superuser:
        messages.error(request, "No tienes permiso para editar mensajes.")
        return redirect("panel_mensajes")

    if request.method == "POST":
        form = MensajeForm(request.POST, instance=mensaje)
        if form.is_valid():
            form.save()
            messages.success(request, "Mensaje actualizado.")
            return redirect("panel_mensajes")
    else:
        form = MensajeForm(instance=mensaje)

    return render(request, "mensajes_admin/editar_mensaje.html", {"form": form})

@login_required
def eliminar_mensaje(request, id):
    mensaje = get_object_or_404(Mensaje, id=id)
    mensaje.delete()
    messages.success(request, "Mensaje eliminado correctamente.")
    return redirect("panel_mensajes")



 
 
# ========= VISTAS EMPLEADO =========

@login_required
def dashboard_empleado(request):
    usuario = request.user
    es_empleado = usuario.groups.filter(name="Empleado").exists()

    # Ignorar cualquier next que apunte a vistas de admin
    next_url = request.GET.get('next')
    if next_url and next_url.startswith('/usuarios_admin/'):
        next_url = None

    campanas_asignadas = CampanaAsignada.objects.filter(empleado=usuario)
    campanas_participadas = CampanaAsignada.objects.filter(
        empleado=usuario,
        realizada=True
    )

    hoy = timezone.now().date()
    campanas_hoy = CampanaAsignada.objects.filter(
        empleado=usuario,
        fecha_asignacion__date=hoy
    )

    calificaciones = Feedback.objects.filter(empleado=usuario)

    eventos = []
    for c in campanas_asignadas:
        eventos.append({
            "title": c.campaña.nombre,
            "start": c.fecha_asignacion.strftime("%Y-%m-%d"),
            "color": "#28a745" if c.realizada else "#007bff",
            "description": f"Estado: {'Realizada' if c.realizada else 'Pendiente'}"
        })

    contexto = {
        'campanas_asignadas': campanas_asignadas,
        'campanas_participadas': campanas_participadas,
        'campanas_hoy': campanas_hoy,
        'calificaciones': calificaciones,
        'es_empleado': es_empleado,
        'campanas_asignadas_json': json.dumps(eventos),
    }

    return render(request, 'dashboard_empleado.html', contexto)

 #--------editar PERFIL EMPLEADO------
 
@login_required
def editar_empleado(request):
    usuario = request.user
    perfil, _ = Perfil.objects.get_or_create(user=usuario)

    if request.method == "POST":
        form = EditarEmpleadoForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario = form.save()
            # Guardar datos de perfil
            perfil.telefono = form.cleaned_data.get('telefono', '')
            perfil.departamento = form.cleaned_data.get('departamento', '')
            perfil.ciudad = form.cleaned_data.get('ciudad', '')
            perfil.direccion = form.cleaned_data.get('direccion', '')
            perfil.save()

            messages.success(request, "Perfil actualizado correctamente.")
            return redirect('dashboard_empleado')
    else:
        form = EditarEmpleadoForm(instance=usuario)

    context = {
        'form': form,
        'usuario': usuario,
        'perfil': perfil,
        'departamentos': ["Cundinamarca", "Antioquia", "Valle del Cauca", "Bogotá"],
        'ciudades_por_departamento': {
            "Cundinamarca": ["Girardot", "Soacha", "Fusagasugá", "Zipaquirá", "Ricaurte"],
            "Antioquia": ["Medellín", "Envigado", "Bello", "Itagüí"],
            "Valle del Cauca": ["Cali", "Palmira", "Buenaventura", "Tuluá"],
            "Bogotá": ["Bogotá"]
        }
    }
    return render(request, 'empleado/editar_empleado.html', context)

#registro pausas empleado
def registrar_pausa(request, campana_id):
    campaña = get_object_or_404(Campaña, id=campana_id)
    evidencia_existente = EvidenciaCampaña.objects.filter(
        campaña=campaña, empleado=request.user
    ).first()

    # Buscamos si la campaña está asignada al usuario
    asignacion = CampanaAsignada.objects.filter(
        campaña=campaña, empleado=request.user
    ).first()

    if request.method == 'POST':
        form = RegistrarEvidenciaCampañaForm(request.POST, request.FILES)
        if form.is_valid():
            evidencia = form.save(commit=False)
            evidencia.campaña = campaña
            evidencia.empleado = request.user
            evidencia.save()

            # Cambiamos estado a "por_aprobacion"
            campaña.estado = 'por_aprobacion'
            campaña.save()

            messages.success(request, "Evidencia registrada. Esperando aprobación del administrador.")
            return redirect('campanas_asignadas')

    else:
        # Si el empleado solo entra a ver la campaña sin subir evidencia
        if campaña.estado == 'activa':
            campaña.estado = 'pausada'
            campaña.save()

        form = RegistrarEvidenciaCampañaForm()

    context = {
        'campaña': campaña,
        'form': form,
        'evidencia': evidencia_existente,
        'asignacion': asignacion,  # ✅ Se pasa al template
    }
    return render(request, 'empleados/registrar_evidencia_campaña.html', context)

@login_required
def campanias_realizadas_empleado(request):
    evidencias = EvidenciaCampaña.objects.filter(
        empleado=request.user
    ).select_related('campaña', 'campaña__codigo')

    return render(request, 'estadisticas_empleados/campanias_realizadas_empleados.html', {
        'evidencias': evidencias
    })
    
def detalle_campania_realizada(request, campana_id):
    campaña = get_object_or_404(Campaña, id=campana_id)

    # ✅ Obtener la evidencia más reciente de este empleado y campaña
    evidencia = EvidenciaCampaña.objects.filter(
        campaña=campaña,
        empleado=request.user
    ).order_by('-fecha_subida').first()

    # ✅ Obtener la asignación (para observaciones del admin)
    asignacion = CampanaAsignada.objects.filter(
        campaña=campaña,
        empleado=request.user
    ).first()

    # ✅ Forzar que el archivo cambie cada vez que se actualiza (evitar caché)
    evidencia_url = None
    if evidencia and evidencia.archivo:
        evidencia_url = f"{evidencia.archivo.url}?v={evidencia.fecha_subida.timestamp()}"

    context = {
        'campaña': campaña,
        'evidencia': evidencia,
        'evidencia_url': evidencia_url,
        'asignacion': asignacion,
        'estado_actual': campaña.estado,
    }
    return render(request, 'estadisticas_empleados/detalle_campania_realizada.html', context)

# === ASIGNACIÓN DE CAMPAÑAS ===
@user_passes_test(lambda u: u.is_superuser)
def asignar_campaña(request):
    if request.method == 'POST':
        form = CampanaAsignadaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Campaña asignada correctamente.")
            return redirect('listar_asignaciones')
    else:
        form = CampanaAsignadaForm()
    return render(request, 'asignaciones/asignar.html', {'form': form})

@user_passes_test(lambda u: u.is_superuser)
def listar_asignaciones(request):
    asignaciones = CampanaAsignada.objects.select_related('campaña', 'empleado')
    return render(request, 'asignaciones/listar.html', {'asignaciones': asignaciones})

@user_passes_test(lambda u: u.is_superuser)
def eliminar_asignacion(request, id):
    asignacion = get_object_or_404(CampanaAsignada, id=id)
    if request.method == 'POST':
        asignacion.delete()
        messages.warning(request, "Asignación eliminada.")
        return redirect('listar_asignaciones')
    return render(request, 'asignaciones/eliminar.html', {'asignacion': asignacion})

def asignar_usuario_campania(request):
    usuarios = Usuario.objects.all()
    campanias = Campaña.objects.all()

    if request.method == "POST":
        usuario_id = request.POST.get("usuario_id")
        campania_id = request.POST.get("campania_id")

        usuario = get_object_or_404(Usuario, id=usuario_id)
        campania = get_object_or_404(Campaña, id=campania_id)

        CampanaAsignada.objects.get_or_create(
            campaña=campania,
            empleado=usuario
        )

        messages.success(request, f"Usuario {usuario.first_name} asignado a campaña {campania.nombre}")
        return redirect("asignar_usuario_campania")

    return render(request, "asignaciones/asignar.html", {
        "usuarios": usuarios,
        "campanias": campanias
    })
    
    

@login_required
def campanas_asignadas(request):
    # -------------------
    # Base: campañas asignadas al usuario
    # -------------------
    asignadas_base = CampanaAsignada.objects.filter(
        empleado=request.user
    ).select_related('campaña', 'campaña__codigo')

    # -------------------
    # Listas desplegables dinámicas para periodicidad
    # -------------------
    periodicidades_disponibles = CampanaAsignada.objects.filter(
        empleado=request.user
    ).values_list('campaña__periodicidad', flat=True).distinct().order_by('campaña__periodicidad')

    # -------------------
    # FILTROS GET
    # -------------------
    estado = request.GET.get('estado', '').strip()
    periodicidad = request.GET.get('periodicidad', '').strip()
    fecha_desde = request.GET.get('fecha_desde', '').strip()
    fecha_hasta = request.GET.get('fecha_hasta', '').strip()

    # -------------------
    # Aplicar filtros sobre las campañas asignadas
    # -------------------
    asignadas = asignadas_base

    if estado:
        asignadas = asignadas.filter(campaña__estado__iexact=estado)

    if periodicidad:
        asignadas = asignadas.filter(campaña__periodicidad__iexact=periodicidad)

    if fecha_desde:
        fecha_desde_parsed = parse_date(fecha_desde)
        if fecha_desde_parsed:
            asignadas = asignadas.filter(campaña__fecha_creacion__date__gte=fecha_desde_parsed)

    if fecha_hasta:
        fecha_hasta_parsed = parse_date(fecha_hasta)
        if fecha_hasta_parsed:
            asignadas = asignadas.filter(campaña__fecha_creacion__date__lte=fecha_hasta_parsed)

    # -------------------
    # Ajustes para la plantilla
    # -------------------
    for a in asignadas:
        a.nombre_campaña = (
            a.campaña.nombre.strip() if a.campaña.nombre else
            getattr(a.campaña.codigo, 'nombre', '').strip() or
            f"Campaña {a.campaña.codigo.codigo}"
        )

        evidencia = EvidenciaCampaña.objects.filter(
            campaña=a.campaña, empleado=request.user
        ).first()

        if evidencia:
            if a.campaña.estado == 'por_aprobacion':
                a.estado_actual = 'En revisión'
            elif a.campaña.estado == 'finalizada':
                a.estado_actual = 'Finalizada'
            else:
                a.estado_actual = a.campaña.estado
        else:
            a.estado_actual = a.campaña.estado

    # -------------------
    # Render
    # -------------------
    return render(request, 'empleados/campanas_asignadas.html', {
        'asignadas': asignadas,
        'filtros': {
            'estado': estado,
            'periodicidad': periodicidad,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
        },
        'periodicidades_disponibles': periodicidades_disponibles,
    })
    
@login_required
def detalle_campana(request, campaña_id):
    campaña = get_object_or_404(Campaña, id=campaña_id)
    return render(request, 'empleados/detalle_campana.html', {
        'campaña': campaña
    })
    
    
#evidencia campaña
def registrar_evidencia_campaña(request, campaña_id):
    campaña = get_object_or_404(Campaña, id=campaña_id)

    # Buscar la asignación correspondiente al empleado actual y la campaña
    asignacion = CampanaAsignada.objects.filter(
        campaña_id=campaña_id,
        empleado=request.user
    ).first()

    # Verificar si ya hay evidencia registrada
    evidencia_existente = EvidenciaCampaña.objects.filter(
        campaña=campaña,
        empleado=request.user
    ).first()

    # ✅ Si la campaña está rechazada, permitimos volver a cargar evidencia
    puede_reenviar = campaña.estado == 'rechazada'

    if request.method == 'POST':
        form = RegistrarEvidenciaCampañaForm(request.POST, request.FILES)
        if form.is_valid():
            # Si ya había una evidencia anterior y fue rechazada, se reemplaza
            if evidencia_existente and puede_reenviar:
                evidencia_existente.archivo.delete(save=False)
                evidencia_existente.archivo = form.cleaned_data['archivo']
                evidencia_existente.save()
            else:
                evidencia = form.save(commit=False)
                evidencia.empleado = request.user
                evidencia.campaña = campaña
                evidencia.save()

            # Cambiar estado a “por_aprobacion”
            campaña.estado = 'por_aprobacion'
            campaña.save()

            # =====================================================
            #   🔔 CREAR NOTIFICACIÓN PARA EL ADMINISTRADOR
            # =====================================================

            from .models import Notificacion, Rol, Usuario  # asegúrate de que esté arriba también

            try:
                # Buscar rol "Admin"
                rol_admin = Rol.objects.get(nombre="Administrador")
                admins = Usuario.objects.filter(rol=rol_admin)
            except Rol.DoesNotExist:
                # Si no existe, enviar a superusuarios
                admins = Usuario.objects.filter(is_superuser=True)

            for admin in admins:
                Notificacion.objects.create(
                    usuario=admin,             # Destinatario
                    remitente=request.user,    # Empleado que envía evidencia
                    campaña=campaña,
                    titulo="📄 Nueva evidencia enviada",
                    mensaje=(
                        f"El empleado {request.user.first_name} {request.user.last_name} "
                        f"ha enviado evidencia para la campaña '{campaña.nombre}'."
                    ),
                    tipo="web"
                )

            # =====================================================

            messages.success(request, "Tu nueva evidencia fue enviada para aprobación.")
            return redirect('campanas_asignadas')

    else:
        form = RegistrarEvidenciaCampañaForm()

    return render(request, 'registrar_evidencia_campaña.html', {
        'campaña': campaña,
        'form': form,
        'evidencia': evidencia_existente,
        'asignacion': asignacion,
        'puede_reenviar': puede_reenviar,  
    })

    
#notificaciones
def detalle_notificacion_empleado(request, pk):
    notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)

    # Marcar como leída
    if not notificacion.abierta:
        notificacion.abierta = True
        notificacion.fecha_apertura = timezone.now()
        notificacion.save(update_fields=["abierta", "fecha_apertura"])

    # Diferenciar tipo de notificación
    campaña = notificacion.campaña
    asignacion = None
    es_mensaje = False

    if campaña:
        asignacion = campaña.campanaasignada_set.filter(empleado=request.user).first()
    else:
        # Si no tiene campaña, es un mensaje informativo
        es_mensaje = True

    return render(request, "notificaciones_empleados/detalle_notificacion_empleado.html", {
        "notificacion": notificacion,
        "campaña": campaña,
        "asignacion": asignacion,
        "es_mensaje": es_mensaje
    })
    
    
def calendario_empleado(request):
    usuario = request.user
    campanas_asignadas = CampanaAsignada.objects.filter(empleado=usuario)

    eventos = []
    for c in campanas_asignadas:
        if c.fecha_vencimiento:  # ✅ Usamos la fecha de vencimiento
            eventos.append({
                "title": c.campaña.nombre if c.campaña else "Campaña sin nombre",
                "start": c.fecha_vencimiento.strftime("%Y-%m-%d"),  # ✅ Fecha de vencimiento
                # Ya no necesitamos descripción
            })

    contexto = {
        "campanas_asignadas": campanas_asignadas,
        "campanas_asignadas_json": json.dumps(eventos, cls=DjangoJSONEncoder, ensure_ascii=False)
    }

    return render(request, "empleados/calendario_empleado.html", contexto)



@login_required
def historial_participacion(request):
    # Solo mostrar historial si es empleado
    rol_usuario = getattr(request.user.rol, 'nombre', '')
    if rol_usuario != 'empleado':
        return render(request, 'empleados/historial_participacion.html', {
            'mensaje': "Los administradores no tienen historial de participación en campañas."
        })

    # Obtener evidencias del empleado
    evidencias = EvidenciaCampaña.objects.filter(
        empleado=request.user
    ).select_related('campaña')

    # Conteo de estados
    conteo_estados = Counter()
    for e in evidencias:
        estado = e.campaña.estado if e.campaña else "Sin estado"
        conteo_estados[estado] += 1

    contexto = {
        'campañas': evidencias,
        'conteo_estados_json': json.dumps(conteo_estados),
    }
    return render(request, 'empleados/historial_participacion.html', contexto)


def recursos_sst(request):
    recursos = RecursoSSTEmpleado.objects.filter(visible=True).select_related('recurso').order_by('-recurso__fecha_subida')

    # Filtros por GET
    tipo = request.GET.get('tipo', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')

    if tipo:
        recursos = recursos.filter(recurso__tipo=tipo)

    if desde:
        recursos = recursos.filter(recurso__fecha_subida__date__gte=parse_date(desde))
    if hasta:
        recursos = recursos.filter(recurso__fecha_subida__date__lte=parse_date(hasta))

    # Paginación
    paginator = Paginator(recursos, 6) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'empleados/recursos_sst.html', {
        'page_obj': page_obj,
        'tipo': tipo,
        'desde': desde,
        'hasta': hasta
    })


@login_required
def mensajes_empleado(request):
    mensajes = Mensaje.objects.all().order_by('-creado')

    # Obtener filtros desde GET
    titulo = request.GET.get('titulo', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')

    # Obtener títulos únicos para la lista desplegable
    titulos_disponibles = Mensaje.objects.values_list('titulo', flat=True).distinct()

    # Aplicar filtros
    if titulo:
        mensajes = mensajes.filter(titulo=titulo)

    if fecha_desde:
        mensajes = mensajes.filter(creado__date__gte=fecha_desde)

    if fecha_hasta:
        mensajes = mensajes.filter(creado__date__lte=fecha_hasta)

    # 🔹 Paginación (10 por página)
    paginator = Paginator(mensajes, 10)
    page_number = request.GET.get('page')
    mensajes = paginator.get_page(page_number)

    return render(request, 'empleados/mensajes_empleado.html', {
        'mensajes': mensajes,
        'titulo': titulo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'titulos_disponibles': titulos_disponibles,
    })
    
@login_required
def detalle_mensaje_empleado(request, id):
    mensaje = get_object_or_404(Mensaje, id=id)

    return render(request, "empleados/detalle_mensaje.html", {
        "mensaje": mensaje
    })





@login_required
def evaluaciones_sst(request):
    return render(request, 'empleados/evaluaciones_sst.html')













def feedback_empleado(request):
    calificaciones = Feedback.objects.filter(empleado=request.user)
    return render(request, 'empleados/feedback.html', {'calificaciones': calificaciones})

@login_required
def encuesta_campaña(request, campaña_id):
    asignacion = get_object_or_404(CampanaAsignada, campaña_id=campaña_id, empleado=request.user)
    if request.method == 'POST':
        form = EncuestaForm(request.POST, request.FILES)
        if form.is_valid():
            encuesta = form.save(commit=False)
            encuesta.campaña_asignada = asignacion
            encuesta.save()
            messages.success(request, 'Encuesta enviada con éxito.')
            return redirect('dashboard_empleado')
    else:
        form = EncuestaForm()
    return render(request, 'encuesta_campaña.html', {'form': form, 'campaña': asignacion.campaña})

@login_required
def feedback_campaña(request, campaña_id):
    asignacion = get_object_or_404(CampanaAsignada, campaña_id=campaña_id, empleado=request.user)
    if request.method == 'POST':
        form = FeedbackForm(request.POST, request.FILES)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.campaña_asignada = asignacion
            feedback.save()
            messages.success(request, 'Gracias por tu calificación.')
            return redirect('dashboard_empleado')
    else:
        form = FeedbackForm()
    return render(request, 'feedback_campaña.html', {'form': form, 'campaña': asignacion.campaña})



# =======================
# VISTAS DE ENCUESTA
# =======================

def encuesta_campaña(request, campaña_id):
    campaña = get_object_or_404(Campaña, id=campaña_id)
    asignacion = get_object_or_404(CampanaAsignada, campaña=campaña, empleado=request.user)

    if request.method == 'POST':
        form = EncuestaForm(request.POST, request.FILES)
        if form.is_valid():
            encuesta = form.save(commit=False)
            encuesta.empleado = request.user
            encuesta.campaña = campaña
            encuesta.fecha = timezone.now()
            encuesta.save()

            # marcar como realizada
            asignacion.realizada = True
            asignacion.save()

            messages.success(request, "¡Encuesta enviada y campaña marcada como realizada!")
            return redirect('dashboard_empleado')
    else:
        form = EncuestaForm()

    return render(request, 'empleados/encuesta_campaña.html', {
        'form': form,
        'campaña': campaña
    })

# =======================
# VISTAS DE FEEDBACK
# =======================

@login_required
def dar_feedback(request, asignacion_id):
    asignacion = get_object_or_404(CampanaAsignada, id=asignacion_id, empleado=request.user)

    if request.method == 'POST':
        form = FeedbackForm(request.POST)
        if form.is_valid():
            feedback = form.save(commit=False)
            feedback.empleado = request.user
            feedback.campaña = asignacion.campaña
            feedback.save()
            messages.success(request, "Gracias por tu feedback.")
            return redirect('dashboard_empleado')
    else:
        form = FeedbackForm()

    return render(request, 'empleados/dar_feedback.html', {
        'form': form,
        'asignacion': asignacion
    })


@user_passes_test(lambda u: u.is_superuser)
def listar_feedback(request):
    feedbacks = Feedback.objects.select_related('empleado', 'campaña').all()
    return render(request, 'feedback/listar_feedback.html', {'feedbacks': feedbacks})

#pausa activa
def ejecutar_pausa(request, pausa_id):
    pausa = get_object_or_404(PausaActiva, pk=pausa_id)
    return render(request, "Notificaciones_empleados/ejecutar_pausa.html", {
        "pausa": pausa
    })
    

